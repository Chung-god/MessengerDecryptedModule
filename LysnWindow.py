from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

from openpyxl.styles import Font, Border, Side, Alignment
import openpyxl
from xml.etree.ElementTree import parse

from exportDB import lysn_userDB, lysn_talkDB
from button import Button
from videoWindow import video, image

import os

class LysnScreen(QDialog):
    def __init__(self, phoneNo):
        super().__init__()
        # 초기화
        self.tableWidget = QTableWidget()
        self.on_off, self.f_name = 0, ''
        self.userColnames, self.userRowlists, self.talkColnames, self.talkRowlists = [], [], [], []

        self.phoneNo = phoneNo
        self.path = f'C:/AppData/{self.phoneNo}/Lysn/'
        self.lysnData()  # 미리 Lysn 데이터 모두 가져오기
        self.setupUI()

    def setupUI(self):

        # Window Backgrond
        palette = QPalette()
        palette.setColor(QPalette.Background, QColor(255, 255, 255))
        self.setAutoFillBackground(True)
        self.setPalette(palette)

        # Window Setting
        self.setGeometry(500, 70, 1200, 800)
        self.setWindowTitle("main")
        self.setFixedSize(self.rect().size())
        self.setContentsMargins(10, 10, 10, 10)

        # Ctrl+ F
        self.shortcut = QShortcut(QKeySequence('Ctrl+f'), self)
        self.shortcut.activated.connect(self.handleFind)

        # back/search button
        self.backButton = Button(QPixmap("image/back.png"), 45, self.showAppWindow)
        self.searchButton = Button(QPixmap("image/search.png"), 45, self.search_items)
        
        # 마우스 커서를 버튼 위에 올리면 모양 바꾸기
        self.backButton.setCursor(QCursor(Qt.PointingHandCursor))
        self.searchButton.setCursor(QCursor(Qt.PointingHandCursor))

        # search text
        self.searchBox = QtWidgets.QLineEdit()
        self.searchBox.setMinimumSize(QtCore.QSize(0, 15))
        self.searchBox.returnPressed.connect(self.search_items)

        # combo box talk
        self.talkComboBox = QComboBox()
        self.talkComboBox.addItem("chats")
        self.talkComboBox.addItem("rooms")
        self.talkComboBox.addItem("lastindex")
        self.talkComboBox.addItem("sqlite_sequence")
        self.talkComboBox.setFixedWidth(100)
        self.talkComboBox.activated.connect(self.talkComboEvent)
        self.talkComboBox.hide()

        # excel button
        self.excelSaveButton = QPushButton(default=False, autoDefault=False)
        self.excelSaveButton.setFixedWidth(100)
        self.excelSaveButton.setText('Save as xls')
        self.excelSaveButton.clicked.connect(self.excelButtonClicked)


        # open combo box
        self.openComboBox = QComboBox()
        self.openComboBox.addItem("user.db")
        self.openComboBox.addItem("talk.db")
        self.openComboBox.setFixedWidth(100)
        self.openComboBox.activated.connect(self.DBClicked)

        # combo box user
        self.userComboBox = QComboBox()
        self.userComboBox.addItem("users")
        self.userComboBox.addItem("sqlite_sequence")
        self.userComboBox.setFixedWidth(100)
        self.userComboBox.activated.connect(self.userComboEvent)

        # close Button
        # btn.clicked.connect(self.close)

        hbox1 = QHBoxLayout()
        hbox1.addWidget(self.backButton)
        hbox1.addStretch(1)
        hbox1.addWidget(self.searchBox)
        hbox1.addWidget(self.searchButton)
        hbox1.addWidget(self.openComboBox)
        hbox2 = QHBoxLayout()
        hbox2.addWidget(self.userComboBox)
        hbox2.addWidget(self.talkComboBox)
        hbox2.addStretch(1)
        hbox2.addWidget(self.excelSaveButton)


        layout = QVBoxLayout()
        layout.addLayout(hbox1)
        layout.addLayout(hbox2)
        layout.addWidget(self.tableWidget)

        self.setLayout(layout)
        self.center()
        self.show()

    def showAppWindow(self):
        self.close()

    # ctrl + f
    def handleFind(self):
        self.on_off = 1
        findDialog = QDialog()
        grid = QGridLayout()
        findDialog.setLayout(grid)
        findLabel = QLabel("Search...", findDialog)
        grid.addWidget(findLabel, 1, 0)
        self.findField = QLineEdit(findDialog)
        grid.addWidget(self.findField, 1, 1)
        findButton = QPushButton("Find", findDialog)
        findButton.clicked.connect(self.search_items)
        grid.addWidget(findButton, 2, 1)
        findDialog.setWindowTitle("Search items")
        findDialog.exec_()
        self.on_off = 0

    # search box
    def search_items(self):

        # rest font
        def reset(self, items):
            for item in items:
                if item == None:
                    pass
                else:
                    item.setBackground(QBrush(Qt.white))
                    item.setForeground(QBrush(Qt.black))
                    item.setFont(QFont())

        if self.on_off == 0:
            text = self.searchBox.text()
            selected_items = self.tableWidget.findItems(self.searchBox.text(), QtCore.Qt.MatchContains)
        else:
            text = self.findField.text()
            selected_items = self.tableWidget.findItems(self.findField.text(), QtCore.Qt.MatchContains)

        allitems = self.tableWidget.findItems("", QtCore.Qt.MatchContains)

        # reset
        reset(self, allitems)

        # highlight the search results
        for item in allitems:
            if item in selected_items:
                item.setBackground(QBrush(Qt.black))
                item.setForeground(QBrush(Qt.white))
                item.setFont(QFont("Helvetica", 9, QFont.Bold))

        if self.searchBox.text() == "" and self.findField.text() != "":
            pass

        elif self.searchBox.text() == "":
            reset(self, allitems)
            print("sb None")
        elif self.on_off == 1 and self.findField.text() == "":
            reset(self, allitems)
            print("ff None")

    # table combo select
    def userComboEvent(self):
        if self.userComboBox.currentText() == 'users':
            colname, rowlist = self.userColnames[0], self.userRowlists[0]
        elif self.userComboBox.currentText() == 'sqlite_sequence':
            colname, rowlist = self.userColnames[1], self.userRowlists[1]

        self.showTable(colname, rowlist)

    def talkComboEvent(self):
        if self.talkComboBox.currentText() == 'chats':
            colname, rowlist = self.talkColnames[0], self.talkRowlists[0]
        elif self.talkComboBox.currentText() == 'rooms':
            colname, rowlist = self.talkColnames[1], self.talkRowlists[1]
        elif self.talkComboBox.currentText() == 'lastindex':
            colname, rowlist = self.talkColnames[2], self.talkRowlists[2]
        elif self.talkComboBox.currentText() == 'sqlite_sequence':
            colname, rowlist = self.talkColnames[3], self.talkRowlists[3]

        self.showTable(colname, rowlist)

    # android id 찾기
    def findAndriodId(self):
        android_id = ''
        tree = parse(self.path + 'settings_secure.xml')
        root = tree.getroot()

        for name in root.iter('setting'):
            d = name.attrib
            for key, value in d.items():
                if key == 'name' and value == 'android_id':
                    android_id = d['value']
        return android_id

    def lysnData(self):
        android_id = self.findAndriodId()
        self.userColnames, self.userRowlists = lysn_userDB(self.path, android_id)
        self.talkColnames, self.talkRowlists = lysn_talkDB(self.path, android_id, self.userRowlists)
        colname, rowlist = self.userColnames[0], self.userRowlists[0]
        self.f_name = "user_db"
        self.showTable(colname, rowlist)

    # DB 버튼 클릭 시
    def DBClicked(self):

        if self.openComboBox.currentText() == 'user.db':
            self.talkComboBox.hide()
            self.userComboBox.show()
            self.userComboBox.setCurrentIndex(0)
            colname, rowlist = self.userColnames[0], self.userRowlists[0]
            self.f_name = "user_db"

        elif self.openComboBox.currentText() == 'talk.db':
            self.userComboBox.hide()
            self.talkComboBox.show()
            self.talkComboBox.setCurrentIndex(0)
            colname, rowlist = self.talkColnames[0], self.talkRowlists[0]
            self.f_name = "talk_db"

        self.showTable(colname, rowlist)

    def tableHeaderClicked(self):
        # 헤더 click 시에만 정렬하고 다시 정렬기능 off
        # 정렬 계속 on 시켜 놓으면 다른 테이블 클릭 시 data 안보이는 현상 발생
        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.setSortingEnabled(False)

    def showTable(self, colname, rowlist):
        self.tableWidget.clear()

        self.tableWidget.setColumnCount(len(colname))  # col 개수 지정
        self.tableWidget.setRowCount(len(rowlist))  # row 개수 지정

        self.tableWidget.setHorizontalHeaderLabels(colname)  # 열 제목 지정
        self.tableHeader = self.tableWidget.horizontalHeader()
        self.tableHeader.sectionClicked.connect(self.tableHeaderClicked)

        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 표 너비 지정
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 표 수정 못하도록
        
        media = -1
        for m in range(len(colname)):
            if list(colname)[m] == '파일':
                media = m
            elif list(colname)[m] == '타입':
                types = m

        # rowlist를 표에 지정하기
        for i in range(len(rowlist)):
            for j in range(len(rowlist[i])):
                if j == media:
                    tp = rowlist[i][types]
                    if tp == 'image': # 사진
                        thumbnail = os.path.splitext(rowlist[i][j])[0].replace("i_o_", "i_t_")
                        mpath = self.path+'LysnMedia/'+thumbnail

                        if os.path.isfile(mpath) == False: # 썸네일 없을 경우
                            if os.path.isfile(self.path+'LysnMedia/'+rowlist[i][j]) == False: #원본파일 없을 경우
                                mpath = 'image/noimage.png'
                            else: # 원본 파일 있는 경우
                                mpath = 'image/image.png'

                        self.btn1 = Button(QPixmap(mpath), 30, self.imageWindow)
                        self.btn1.setText(rowlist[i][j])
                        self.tableWidget.setCellWidget(i,j,self.btn1)

                    elif tp == 'video': # 비디오
                        thumbnail = os.path.splitext(rowlist[i][j])[0].replace("v_o_", "v_t_")
                        mpath = self.path+'LysnMedia/'+thumbnail

                        if os.path.isfile(mpath) == False:
                            if os.path.isfile(self.path+'LysnMedia/'+rowlist[i][j]) == False:
                                mpath = 'image/noimage.png'
                            else:
                                mpath = 'image/video.png'
                            
                        self.btn2 = Button(QPixmap(mpath), 30, self.videoWindow)
                        self.btn2.setText(rowlist[i][j])
                        self.tableWidget.setCellWidget(i,j,self.btn2)
                else: # 텍스트
                    item = QTableWidgetItem(str(rowlist[i][j]))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget.setItem(i, j, item)
        self.tableWidget.verticalHeader().setDefaultSectionSize(130)
        self.colname = colname
        self.rowlist = rowlist

    def imageWindow(self):
        file = self.sender().text()
        mediaPath = self.path+'LysnMedia/'+file
        image(mediaPath)
        
    def videoWindow(self):
        file = self.sender().text()
        mediaPath = self.path+'LysnMedia/'+file
        video(mediaPath)
        
    def center(self):
        frame_info = self.frameGeometry()
        display_center = QDesktopWidget().availableGeometry().center()
        frame_info.moveCenter(display_center)
        self.move(frame_info.topLeft())

    # 엑셀 추출 버튼 클릭 시
    def excelButtonClicked(self):
        if self.f_name == '':
            return

        # create Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Lysn"
        col_excel = list(self.colname)[0:5]

        for x in range(1, len(col_excel) + 1):
            sheet.cell(row=1, column=x).value = col_excel[x - 1]

        for x in range(0, len(self.rowlist)):
            for y in range(1, len(self.rowlist[x]) + 1):
                sheet.cell(row=x + 2, column=y).value = str(self.rowlist[x][y - 1])

        # resize the cell
        for x in range(0, len(self.colname)):
            MAX = 1
            for y in range(1, len(self.rowlist) + 1):
                cell_size = len(str(self.rowlist[y - 1][x]))
                if MAX < cell_size:
                    MAX = cell_size
                    sheet.column_dimensions[chr(65 + x)].width = MAX + 5
                sheet.row_dimensions[y].height = 20
        sheet.row_dimensions[y + 1].height = 20

        # change the font
        for x in range(1, len(self.colname) + 1):
            cell = sheet[chr(64 + x) + "1"]
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(right=Side(border_style="thick"), bottom=Side(border_style="thick"))

        for x in range(len(self.rowlist)):
            for y in range(len(self.rowlist[x])):
                cell = sheet[chr(65 + y) + str(x + 2)]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(right=Side(border_style="thick"))

        wb.save("Lysn_" + self.f_name + ".xlsx")


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion')) # --> 없으면, 헤더색 변경 안됨.
    phoneNo = 'SM-G955N'
    ui = LysnScreen(phoneNo)
    sys.exit(app.exec_())
