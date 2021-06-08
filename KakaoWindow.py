from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

from openpyxl.styles import Font, Border, Side, Alignment
import openpyxl
from xml.etree.ElementTree import parse

from exportDB import *
from button import Button


class KakaoScreen(QDialog):
    def __init__(self, phoneNo):
        super().__init__()
        # 초기화
        self.tableWidget = QTableWidget()
        self.on_off, self.f_name = 0, ''
        self.kakaoColnames, self.kakaoRowlists, self.kakao2Colnames, self.kakao2Rowlists = [], [], [], []

        self.phoneNo = phoneNo
        self.path = f'/home/ksjr06/바탕화면/05_19/MessengerDecryptedModule--Lysn/GUI/kakao/'
        
        self.kakaoData()  # 미리 Lysn 데이터 모두 가져오기
        self.setupUI()

    def setupUI(self):

        # Window Backgrond
        palette = QPalette()
        palette.setColor(QPalette.Background, QColor(242, 242, 242))
        self.setAutoFillBackground(True)
        self.setPalette(palette)

        # Window Setting
        self.setGeometry(500, 70, 800, 600)
        self.setWindowTitle("main")
        self.setFixedSize(self.rect().size())
        self.setContentsMargins(10, 10, 10, 10)

        # Ctrl+ F
        self.shortcut = QShortcut(QKeySequence('Ctrl+f'), self)
        self.shortcut.activated.connect(self.handleFind)

        # back/search button
        self.backButton = Button(QPixmap("image/back.png"), 35, self.showAppWindow)
        self.searchButton = Button(QPixmap("image/search.png"), 35, self.search_items)
        self.backButton.setStyleSheet('background:transparent')
        self.searchButton.setStyleSheet('background:transparent')

        # 마우스 커서를 버튼 위에 올리면 모양 바꾸기
        self.backButton.setCursor(QCursor(Qt.PointingHandCursor))
        self.searchButton.setCursor(QCursor(Qt.PointingHandCursor))

        # search text
        self.searchBox = QtWidgets.QLineEdit()
        self.searchBox.setMinimumSize(QtCore.QSize(0, 15))

        # combo box name => db2
        self.kakao2ComboBox = QComboBox()
        self.kakao2ComboBox.addItem("friends")
        self.kakao2ComboBox.setFixedWidth(100)
        self.kakao2ComboBox.activated.connect(self.kakao2ComboEvent)
        self.kakao2ComboBox.hide()
        
        # combo box mes => db1
        #self.kakaoComboBox = QComboBox()
        #self.kakaoComboBox.addItem("chat_logs")
        #self.kakaoComboBox.setFixedWidth(100)
        #self.kakaoComboBox.activated.connect(self.kakaoComboEvent)
        

        # excel button
        self.excelSaveButton = QPushButton()
        self.excelSaveButton.setFixedWidth(100)
        self.excelSaveButton.setText('xls')
        self.excelSaveButton.clicked.connect(self.excelButtonClicked)

        # open combo box
        self.openComboBox = QComboBox()
        self.openComboBox.addItem("KakaoTalk.db")
        self.openComboBox.addItem("KakaoTalk2.db")
        self.openComboBox.setFixedWidth(100)
        self.openComboBox.activated.connect(self.DBClicked)
        
        # combo box mes => db1
        self.kakaoComboBox = QComboBox()
        self.kakaoComboBox.addItem("chat_logs")
        self.kakaoComboBox.setFixedWidth(100)
        self.kakaoComboBox.activated.connect(self.kakaoComboEvent)

        # combo box name => db2
        #self.kakao2ComboBox = QComboBox()
        #self.kakao2ComboBox.addItem("friends")
        #self.kakao2ComboBox.setFixedWidth(100)
        #self.kakao2ComboBox.activated.connect(self.kakao2ComboEvent)
        #self.kakao2ComboBox.hide()

        hbox1 = QHBoxLayout()
        hbox1.addWidget(self.backButton)
        hbox1.addStretch(1)
        hbox1.addWidget(self.searchBox)
        hbox1.addWidget(self.searchButton)
        hbox1.addWidget(self.openComboBox)
        hbox2 = QHBoxLayout()
        hbox2.addWidget(self.kakaoComboBox)
        hbox2.addWidget(self.kakao2ComboBox)
        hbox2.addStretch(1)
        hbox2.addWidget(self.excelSaveButton)

        layout = QVBoxLayout()
        layout.addLayout(hbox1)
        layout.addLayout(hbox2)
        layout.addWidget(self.tableWidget)

        self.setLayout(layout)
        self.center()
        self.show()

    def showAppWindow(self):
        self.close()

    # ctrl + f
    def handleFind(self):
        self.on_off = 1
        findDialog = QDialog()
        grid = QGridLayout()
        findDialog.setLayout(grid)
        findLabel = QLabel("Search...", findDialog)
        grid.addWidget(findLabel, 1, 0)
        self.findField = QLineEdit(findDialog)
        grid.addWidget(self.findField, 1, 1)
        findButton = QPushButton("Find", findDialog)
        findButton.clicked.connect(self.search_items)
        grid.addWidget(findButton, 2, 1)
        findDialog.setWindowTitle("Search items")
        findDialog.exec_()
        self.on_off = 0

    # search box
    def search_items(self):

        # rest font
        def reset(self, items):
            for item in items:
                item.setBackground(QBrush(Qt.white))
                item.setForeground(QBrush(Qt.black))
                item.setFont(QFont())

        if self.on_off == 0:
            text = self.searchBox.text()
            selected_items = self.tableWidget.findItems(self.searchBox.text(), QtCore.Qt.MatchContains)
        else:
            text = self.findField.text()
            selected_items = self.tableWidget.findItems(self.findField.text(), QtCore.Qt.MatchContains)

        allitems = self.tableWidget.findItems("", QtCore.Qt.MatchContains)

        # reset
        reset(self, allitems)

        # highlight the search results
        for item in allitems:
            if item in selected_items:
                item.setBackground(QBrush(Qt.black))
                item.setForeground(QBrush(Qt.white))
                item.setFont(QFont("Helvetica", 11, QFont.Bold))

        if self.searchBox.text() == "":
            reset(self, allitems)
            print("sb None")
        elif self.on_off == 1 and self.findField.text() == "":
            reset(self, allitems)
            print("ff None")

    # table combo select
    def kakaoComboEvent(self):
        if self.kakaoComboBox.currentText() == 'chat_logs':
            colname, rowlist = self.kakaoColnames[0], self.kakaoRowlists[0]

        self.showTable(colname, rowlist)

    def kakao2ComboEvent(self):
        if self.kakao2ComboBox.currentText() == 'friends':
            colname, rowlist = self.kakao2Colnames[0], self.kakao2Rowlists[0]

        self.showTable(colname, rowlist)

    # android id 찾기
#    def findAndriodId(self):
#        android_id = ''
#
#        tree = parse(self.path + 'settings_secure.xml')
#        root = tree.getroot()

#       for name in root.iter('setting'):
#            d = name.attrib
#            for key, value in d.items():
#                if key == 'name' and value == 'android_id':
#                    android_id = d['value']
#        print(android_id)
#        return android_id
#    def findAndriodId(self):
#    	android_id = '073ba681a84bea93'
#    	return android_id
    
    	
    def kakaoData(self):
        self.kakaoColnames, self.kakaoRowlists = KaKaoTalk_DB_1(self.path)
        self.kakao2Colnames, self.kakao2Rowlists = KaKaoTalk_DB_2(self.path)
        colname, rowlist = self.kakaoColnames[0], self.kakaoRowlists[0]
        self.f_name = "KakaoTalk_db" #이게 뭘 뜻하는지 찾아볼 것
        self.showTable(colname, rowlist)

    # DB 버튼 클릭 시
    def DBClicked(self):

        if self.openComboBox.currentText() == 'KakaoTalk.db':
            self.kakao2ComboBox.hide()
            self.kakaoComboBox.show()
            self.kakaoComboBox.setCurrentIndex(0)
            colname, rowlist = self.kakaoColnames[0], self.kakaoRowlists[0]
            self.f_name = "KakaoTalk_db"

        elif self.openComboBox.currentText() == 'KakaoTalk2.db':
            self.kakaoComboBox.hide()
            self.kakao2ComboBox.show()
            self.kakao2ComboBox.setCurrentIndex(0)
            colname, rowlist = self.kakao2Colnames[0], self.kakao2Rowlists[0]
            self.f_name = "KakaoTalk2_db"

        self.showTable(colname, rowlist)

    def tableHeaderClicked(self):
        # 헤더 click 시에만 정렬하고 다시 정렬기능 off
        # 정렬 계속 on 시켜 놓으면 다른 테이블 클릭 시 data 안보이는 현상 발생
        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.setSortingEnabled(False)

    def showTable(self, colname, rowlist):
        self.tableWidget.clear()

        self.tableWidget.setColumnCount(len(colname))  # col 개수 지정
        self.tableWidget.setRowCount(len(rowlist))  # row 개수 지정

        self.tableWidget.setHorizontalHeaderLabels(colname)  # 열 제목 지정
        self.tableHeader = self.tableWidget.horizontalHeader()
        self.tableHeader.sectionClicked.connect(self.tableHeaderClicked)

        #self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)  # 표 너비 지정
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 표 수정 못하도록

        media = -1
        for m in range(len(colname)):
            if list(colname)[m] == '파일':
                media = m

        # rowlist를 표에 지정하기
        for i in range(len(rowlist)):
            for j in range(len(rowlist[i])):
                if j == media:
                    item = self.getImageLabel(rowlist[i][j])
                    self.tableWidget.setCellWidget(i, j, item)
                else:
                    item = QTableWidgetItem(str(rowlist[i][j]))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget.setItem(i, j, item)
        self.tableWidget.verticalHeader().setDefaultSectionSize(80)
        self.colname = colname
        self.rowlist = rowlist

    def getImageLabel(self, image):
        imageLabel = QLabel()
        imageLabel.setScaledContents(True)
        pixmap = QPixmap()
        pixmap.loadFromData(image, 'jpg')
        imageLabel.setPixmap(pixmap)
        return imageLabel

    def center(self):
        frame_info = self.frameGeometry()
        display_center = QDesktopWidget().availableGeometry().center()
        frame_info.moveCenter(display_center)
        self.move(frame_info.topLeft())

    # 엑셀 추출 버튼 클릭 시
    def excelButtonClicked(self):
        if self.f_name == '':
            return

        # create Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Lysn"
        col_excel = list(self.colname)[0:5]

        for x in range(1, len(col_excel) + 1):
            sheet.cell(row=1, column=x).value = col_excel[x - 1]

        for x in range(0, len(self.rowlist)):
            for y in range(1, len(self.rowlist[x]) + 1):
                sheet.cell(row=x + 2, column=y).value = str(self.rowlist[x][y - 1])

        # resize the cell
        for x in range(0, len(self.colname)):
            MAX = 1
            for y in range(1, len(self.rowlist) + 1):
                cell_size = len(str(self.rowlist[y - 1][x]))
                if MAX < cell_size:
                    MAX = cell_size
                    sheet.column_dimensions[chr(65 + x)].width = MAX + 5
                sheet.row_dimensions[y].height = 20
        sheet.row_dimensions[y + 1].height = 20

        # change the font
        for x in range(1, len(self.colname) + 1):
            cell = sheet[chr(64 + x) + "1"]
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(right=Side(border_style="thick"), bottom=Side(border_style="thick"))

        for x in range(len(self.rowlist)):
            for y in range(len(self.rowlist[x])):
                cell = sheet[chr(65 + y) + str(x + 2)]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(right=Side(border_style="thick"))

        wb.save("Lysn_" + self.f_name + ".xlsx")


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    phoneNo = 'SM-N976N'
    ui = LysnScreen(phoneNo)
    sys.exit(app.exec_())