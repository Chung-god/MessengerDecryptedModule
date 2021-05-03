import sys
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtGui import *
from lysn import lysn_userDB, lysn_talkDB
from openpyxl.styles import Font, Border, Side, Alignment
import openpyxl

class DB(QWidget):
    def __init__(self):
        super().__init__()
        self.table = QTableWidget()
        self.initUI()
        
        
    def initUI(self):
        self.f_name = ''
        self.colname = []
        self.rowlist = []
        
    
        # 앱 버튼 생성
        kakaoButton = QPushButton("KakaoTalk")
        wechatButton = QPushButton("WeChat")
        lysnButton = QPushButton("Lysn")
        tongButton = QPushButton("TongTong")
        self.searchBox = QtWidgets.QLineEdit("Search...")
        self.searchButton = QtWidgets.QPushButton("Search")
        
        # Ctrl+ F 
        shortcut = QShortcut(QKeySequence('Ctrl+f'), self)
        shortcut.activated.connect(self.handleFind)
     
        # 버튼 클릭 시 실행되는 함수 지정
        lysnButton.clicked.connect(self.lysnButtonClicked)

        # android id 라벨, 텍스트 초기화
        # 처음에 숨겨놨다가 Lysn 버튼 누르면 보여주기
        self.androidIdLabel = QLabel('Android Id : ')
        self.androidIdLabel.setFixedWidth(100)
        self.androidIdLabel.hide() # 숨기기
        self.lysnAndroidId = QLineEdit("4f77d977f3f1c488")
        self.lysnAndroidId.setFixedWidth(300)
        self.lysnAndroidId.hide()

        # user.db와 talk.db 버튼 생성
        self.lysnButton_user = QPushButton("User.db")
        self.lysnButton_user.hide()
        self.lysnButton_talk = QPushButton("Talk.db")
        self.lysnButton_talk.hide()

        # 엑셀 추출 버튼 생성
        self.excelButton = QPushButton("Excel Export")
        self.excelButton.clicked.connect(self.excelButtonClicked)

        # 수평 레이아웃 생성
        hbox1 = QHBoxLayout()
        hbox2 = QHBoxLayout()
        hbox3 = QHBoxLayout()
        hbox4 = QHBoxLayout()
        hbox5 = QHBoxLayout()
        hbox6 = QHBoxLayout()

        hbox1.addWidget(kakaoButton)
        hbox2.addWidget(wechatButton)

        hbox3.addWidget(lysnButton)
        hbox4.addWidget(self.androidIdLabel)
        hbox4.addWidget(self.lysnAndroidId)   
        hbox4.addWidget(self.lysnButton_user)
        hbox4.addWidget(self.lysnButton_talk)

        hbox5.addWidget(tongButton)
        hbox5.addWidget(self.searchBox)
        hbox5.addWidget(self.searchButton)
        
        hbox6.addWidget(self.excelButton)
        
        # 수직 레이아웃 생성
        layout = QVBoxLayout()
        layout.addLayout(hbox1)
        layout.addLayout(hbox2)
        layout.addLayout(hbox3)
        layout.addLayout(hbox4)
        layout.addLayout(hbox5)
        layout.addWidget(self.table)
        layout.addLayout(hbox6)
        
        
        self.on_off = 0
        self.searchButton.clicked.connect(self.search_items) # search
        self.setLayout(layout) # 레이아웃 설정
        self.setGeometry(1000,1000,1000,1000) # window 화면 크기
        self.setWindowTitle('messenger app artifact') # window 이름 지정
        self.center()
        self.show()
    
    
    # search box
    def search_items(self):
        
        #rest font
        def reset(self, items):
            for item in items:
                item.setBackground(QBrush(Qt.white))
                item.setForeground(QBrush(Qt.black))
                item.setFont(QFont())
            
        
        if self.on_off == 0:
                text = self.searchBox.text()
                selected_items = self.table.findItems(self.searchBox.text(), QtCore.Qt.MatchContains)
        else:
                text = self.findField.text()
                selected_items = self.table.findItems(self.findField.text(), QtCore.Qt.MatchContains)
            
            
        allitems = self.table.findItems("", QtCore.Qt.MatchContains)
        
        # reset
        reset(self, allitems)
             
        # highlight the search results
        for item in allitems:
            if item in selected_items:   
                item.setBackground(QBrush(Qt.black))
                item.setForeground(QBrush(Qt.white))
                item.setFont(QFont("Helvetica", 11, QFont.Bold))
               
        if self.searchBox.text() == "":
            reset(self, allitems)
            print("sb None")
        elif self.on_off == 1 and self.findField.text() =="":
            reset(self, allitems)
            print("ff None")
        
   
    # ctrl + f
    def handleFind(self):
        self.on_off = 1
        findDialog = QDialog()
        grid = QGridLayout()
        findDialog.setLayout(grid)
        findLabel = QLabel("Search...", findDialog)
        grid.addWidget(findLabel,1,0)
        self.findField = QLineEdit(findDialog)
        grid.addWidget(self.findField,1,1)
        findButton = QPushButton("Find", findDialog)
        findButton.clicked.connect(self.search_items)
        grid.addWidget(findButton,2,1)
        findDialog.setWindowTitle("Search items")
        findDialog.exec_()
        self.on_off = 0
        

    def center(self):
        frame_info = self.frameGeometry()
        display_center = QDesktopWidget().availableGeometry().center()
        frame_info.moveCenter(display_center)
        self.move(frame_info.topLeft())

    # Lysn 버튼 클릭 시
    def lysnButtonClicked(self):
        # user.db 버튼이 보여져 있다면 숨기도록, 숨겨져 있다면 보이도록 바꾼다.
        if self.lysnButton_user.isVisible() == True:
            self.lysnButton_user.hide()
            self.lysnButton_talk.hide()
            self.androidIdLabel.hide()
            self.lysnAndroidId.hide()
        else:
            self.lysnButton_user.show()
            self.lysnButton_talk.show()
            self.androidIdLabel.show()
            self.lysnAndroidId.show()

            # DBCliced 함수와 연결
            self.lysnButton_user.clicked.connect(self.DBClicked) 
            self.lysnButton_talk.clicked.connect(self.DBClicked)

    # DB 버튼 클릭 시
    def DBClicked(self):
        
        # android_id 입력 안하면 함수 실행 안하도록
        # android_id = '4f77d977f3f1c488'
        android_id = self.lysnAndroidId.text()
        if android_id == '':
            return
        
        # db file 선택해서 경로 받아오기
        global filename
        filename = QFileDialog.getOpenFileName(self, 'Open File')

        # db file 경로에 따라 db안에 있는 아티팩트 가져오기
        # colname에 열 제목 담기, rowlist에 각 행마다 리스트로 담기
        if filename[0][-7:] == 'user.db':
            self.colname, self.rowlist = lysn_userDB(filename[0], android_id)
            self.f_name = "user_db"

        elif filename[0][-7:] == 'talk.db':
            self.colname, self.rowlist = lysn_talkDB(filename[0], android_id)
            self.f_name = "talk_db"
        
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers) # 표 수정 못하도록
        self.table.setColumnCount(len(self.colname)) # col 개수 지정
        self.table.setRowCount(len(self.rowlist)) # row 개수 지정
        
        self.table.setHorizontalHeaderLabels(self.colname) # 열 제목 지정
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # 표 너비 지정
        # rowlist를 표에 지정하기
        for i in range(len(self.rowlist)):
            for j in range(len(self.rowlist[i])):
                self.table.setItem(i, j, QTableWidgetItem(str(self.rowlist[i][j]))) 
            
    # 엑셀 추출 버튼 클릭 시
    def excelButtonClicked(self):
        if self.f_name == '':
            return

        #create Excel    
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Lysn"
        col_excel = list(self.colname)[0:5]
        
        for x in range(1, len(col_excel) + 1):
            sheet.cell(row = 1, column = x).value = col_excel[x - 1]
            
        for x in range(0, len(self.rowlist)):
            for y in range(1, len(self.rowlist[x]) + 1):
                sheet.cell(row = x + 2, column = y).value = str(self.rowlist[x][y - 1])
                
        #resize the cell
        for x in range(0, len(self.colname)):
            MAX = 1
            for y in range(1, len(self.rowlist) + 1):
                cell_size = len(str(self.rowlist[y - 1][x]))
                if MAX < cell_size:
                    MAX = cell_size
                    sheet.column_dimensions[chr(65 + x)].width = MAX + 5
                sheet.row_dimensions[y].height = 20
        sheet.row_dimensions[y + 1].height = 20
        
        #change the font 
        for x in range(1, len(self.colname) + 1):
            cell = sheet[chr(64 + x) + "1"]
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(right=Side(border_style="thick"), bottom=Side(border_style="thick"))

        for x in range(len(self.rowlist)):
            for y in range(len(self.rowlist[x])):
                cell = sheet[chr(65 + y) + str(x + 2)]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(right=Side(border_style="thick"))
 
        wb.save("Lysn_" + self.f_name + ".xlsx")
        

        
app = QApplication(sys.argv)
db = DB()
db.show()
sys.exit(app.exec_())
