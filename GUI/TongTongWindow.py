from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

from openpyxl.styles import Font, Border, Side, Alignment
import openpyxl
from xml.etree.ElementTree import parse

from exportDB import tongtong_gcmDB
from button import Button
from videoWindow import video, image

import os
import copy

class TongTongScreen(QDialog):
    def __init__(self, path):
        super().__init__()
        # 초기화
        self.setWindowFlags(Qt.WindowTitleHint | Qt.WindowCloseButtonHint )
        self.tableWidget = QTableWidget()
        self.on_off, self.f_name = 0, ''
        self.gcmColnames, self.gcmRowlists = [], []

        self.path = path
        self.TongTongData() # 미리 TongTong 데이터 모두 가져오기
        self.setupUI()
        

    def setupUI(self):
        
        # Window Backgrond
        palette = QPalette()
        palette.setColor( QPalette.Background , QColor(255, 255, 255))
        self.setAutoFillBackground(True)
        self.setPalette(palette)
        
        # Window Setting
        self.setGeometry(500, 70, 1200, 800)
        self.setWindowTitle("main")
        self.setFixedSize(self.rect().size())
        self.setContentsMargins(10,10,10,10)

        # Ctrl+ F 
        self.shortcut = QShortcut(QKeySequence('Ctrl+f'), self)
        self.shortcut.activated.connect(self.handleFind)

        # back/search button
        self.backButton = Button(QPixmap("image/back.png"), 35, self.showAppWindow)
        self.searchButton = Button(QPixmap("image/search.png"), 35, self.search_items)
        self.backButton.setStyleSheet('background:transparent')
        self.searchButton.setStyleSheet('background:transparent')

        # 마우스 커서를 버튼 위에 올리면 모양 바꾸기
        self.backButton.setCursor(QCursor(Qt.PointingHandCursor))
        self.searchButton.setCursor(QCursor(Qt.PointingHandCursor))

        self.backButton.setToolTip('뒤로가기')
        self.searchButton.setToolTip('찾기 버튼\n단축키 : ctrl + f')

        # search text
        self.searchBox = QtWidgets.QLineEdit()
        self.searchBox.setMinimumSize(QtCore.QSize(0, 15))
        self.searchBox.returnPressed.connect(self.search_items)

        # excel button
        self.excelSaveButton = QPushButton(default=False, autoDefault=False)
        self.excelSaveButton.setFixedWidth(100)
        self.excelSaveButton.setText('Save as xls')
        self.excelSaveButton.clicked.connect(self.excelButtonClicked)
        self.excelSaveButton.setToolTip('현재 보고있는 표를 엑셀로 저장하는 버튼')

        # combo box gcm
        self.gcmComboBox = QComboBox()
        self.gcmComboBox.addItem("chatting")
        self.gcmComboBox.addItem("chatRoomList")
        self.gcmComboBox.addItem("friendsList")
        self.gcmComboBox.setFixedWidth(100)
        self.gcmComboBox.activated.connect(self.gcmComboEvent)
        self.gcmComboBox.setToolTip('gcm.db의 tables')

        # combo chat room
        self.chatRoomComboBox = QComboBox()
        for i in range(self.chatRoomLen):
            self.chatRoomComboBox.addItem(self.chatRoomName[i])
        self.chatRoomComboBox.setFixedWidth(100)
        self.chatRoomComboBox.activated.connect(self.chatRoomComboEvent)
        self.chatRoomComboBox.setToolTip('Chat Room')
        
        # open combo box
        self.openComboBox = QComboBox()
        self.openComboBox.addItem("gcm.db")
        self.openComboBox.setFixedWidth(100)
        self.openComboBox.setToolTip('TongTong 데이터베이스')

        hbox1 = QHBoxLayout()
        hbox1.addWidget(self.backButton)
        hbox1.addStretch(1)
        hbox1.addWidget(self.searchBox)
        hbox1.addWidget(self.searchButton)
        hbox1.addWidget(self.openComboBox)
        hbox2 = QHBoxLayout()
        hbox2.addWidget(self.gcmComboBox)
        hbox2.addWidget(self.chatRoomComboBox)
        hbox2.addStretch(1)
        hbox2.addWidget(self.excelSaveButton)
        
        layout = QVBoxLayout()
        layout.addLayout(hbox1)
        layout.addLayout(hbox2)
        layout.addWidget(self.tableWidget)

        self.setLayout(layout)
        self.center()
        self.show()

    def showAppWindow(self):
        self.close()

    # ctrl + f
    def handleFind(self):
        self.on_off = 1
        findDialog = QDialog()
        grid = QGridLayout()
        findDialog.setLayout(grid)
        findLabel = QLabel("Search...", findDialog)
        grid.addWidget(findLabel,1,0)
        self.findField = QLineEdit(findDialog)
        grid.addWidget(self.findField,1,1)
        findButton = QPushButton("Find", findDialog)
        findButton.clicked.connect(self.search_items)
        grid.addWidget(findButton,2,1)
        findDialog.setWindowTitle("Search items")
        findDialog.exec_()
        self.on_off = 0
    
    # search box
    def search_items(self):

        # rest font
        def reset(self, items):
            for item in items:
                if item == None:
                    pass
                else:
                    item.setBackground(QBrush(Qt.white))
                    item.setForeground(QBrush(Qt.black))
                    item.setFont(QFont())

        if self.on_off == 0:
            text = self.searchBox.text()
            selected_items = self.tableWidget.findItems(self.searchBox.text(), QtCore.Qt.MatchContains)
        elif self.on_off == 1:
            text = self.findField.text()
            selected_items = self.tableWidget.findItems(self.findField.text(), QtCore.Qt.MatchContains)

        allitems = self.tableWidget.findItems("", QtCore.Qt.MatchContains)

        # reset
        reset(self, allitems)

        # highlight the search results
        for item in allitems:
            if item in selected_items:
                if item == None:
                    pass
                else:
                    item.setBackground(QBrush(Qt.black))
                    item.setForeground(QBrush(Qt.white))
                    item.setFont(QFont("Helvetica", 9, QFont.Bold))

        if self.searchBox.text() == "" and self.on_off == 0:
            reset(self, allitems)
            pass

        elif self.searchBox.text() == "":
            reset(self, allitems)
            print("sb None")
        elif self.on_off == 1 and self.findField.text() == "":
            reset(self, allitems)
            print("ff None")


    # table combo select
    def gcmComboEvent(self):
        if self.gcmComboBox.currentText() == 'chatting':
            self.chatRoomComboBox.show()
            self.chatRoomComboBox.setCurrentIndex(0)
            colname, rowlist = self.gcmColnames[0], self.chatrowlists[0]
        elif self.gcmComboBox.currentText() == 'chatRoomList':
            self.chatRoomComboBox.hide()
            colname, rowlist = self.gcmColnames[1], self.gcmRowlists[1]
        elif self.gcmComboBox.currentText() == 'friendsList':
            self.chatRoomComboBox.hide()
            colname, rowlist = self.gcmColnames[2], self.gcmRowlists[2]

        self.showTable(colname, rowlist)
    
    def chatroom(self):
        self.gcmColnames[0] = list(self.gcmColnames[0])
        self.gcmColnames[0][2] = '받는사람'
        self.chatrowlists = []
        talkrowlist = self.gcmRowlists[0]

        for k in range(self.chatRoomLen):
            crowlist = []
            for i in range(len(self.gcmRowlists[0])):
                people = copy.deepcopy(self.chatRoomPeople[k])
                if talkrowlist[i][2] == self.chatRoomNum[k]:
                    if talkrowlist[i][1] in people:
                        people.remove(talkrowlist[i][1])
                    people=', '.join(people)
                    talkrowlist[i][2] = people
                    if people == '':
                        talkrowlist[i][2] = talkrowlist[i][1]
                    crowlist.append(talkrowlist[i])
            
            self.chatrowlists.append(crowlist)
    
    def chatRoomComboEvent(self):
        colname = self.gcmColnames[0]
        for k in range(self.chatRoomLen):
            if self.chatRoomComboBox.currentText() == self.chatRoomName[k]:
                rowlist = self.chatrowlists[k]
        
        self.showTable(colname, rowlist)

    def TongTongData(self):
        self.gcmColnames, self.gcmRowlists = tongtong_gcmDB(self.path)

        for i in range(len(self.gcmRowlists[0])):
            del self.gcmRowlists[0][i][1]
        self.gcmColnames[0] = (list(self.gcmColnames[0]))
        del self.gcmColnames[0][1]

        self.chatRoomLen = len(self.gcmRowlists[1])
        self.chatRoomNum = [self.gcmRowlists[1][i][0] for i in range(self.chatRoomLen)]
        self.chatRoomName = [self.gcmRowlists[1][i][1] for i in range(self.chatRoomLen)]
        self.chatRoomPeople = [self.gcmRowlists[1][i][4].split(', ') for i in range(self.chatRoomLen)]

        self.chatroom()
        
        colname, rowlist = self.gcmColnames[0], self.chatrowlists[0]
        self.f_name = "gcm_db"
        self.showTable(colname, rowlist)
    
    def tableHeaderClicked(self):
        # 헤더 click 시에만 정렬하고 다시 정렬기능 off
        # 정렬 계속 on 시켜 놓으면 다른 테이블 클릭 시 data 안보이는 현상 발생
        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.setSortingEnabled(False)
        
    def showTable(self, colname, rowlist):
        self.tableWidget.clear()

        media = -1
        for m in range(len(colname)):
            if list(colname)[m] == '미디어':
                media = m
            if list(colname)[m] == '비디오':
                video = m
            if list(colname)[m] == '타입':
                types = m
        # col 개수 지정
        if media != -1:
            self.tableWidget.setColumnCount(len(colname)-2)
        else:
            self.tableWidget.setColumnCount(len(colname))
        self.tableWidget.setRowCount(len(rowlist)) # row 개수 지정
        
        self.tableWidget.setHorizontalHeaderLabels(colname) # 열 제목 지정
        self.tableHeader = self.tableWidget.horizontalHeader()
        self.tableHeader.sectionClicked.connect(self.tableHeaderClicked)

        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # 표 너비 지정
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers) # 표 수정 못하도록
        
        # rowlist를 표에 지정하기
        for i in range(len(rowlist)):
            for j in range(len(rowlist[i])):
                if j == media: # 사진 or 비디오
                    mpath = rowlist[i][j]
                    vpath = rowlist[i][video]
                    tp = rowlist[i][types]
                    if tp == 'image':
                        self.btn1 = Button(QPixmap(mpath), 30, self.imageWindow)
                        self.btn1.setText(rowlist[i][j])
                        self.tableWidget.setCellWidget(i,j,self.btn1)
                    elif tp == 'video':
                        self.btn1 = Button(QPixmap(mpath), 30, self.videoWindow)
                        self.btn1.setText(vpath)
                        self.tableWidget.setCellWidget(i,j,self.btn1)
                else: # 텍스트
                    item = QTableWidgetItem(str(rowlist[i][j]))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget.setItem(i, j, item)
        
        self.tableWidget.verticalHeader().setDefaultSectionSize(130)
        self.colname = colname
        self.rowlist = rowlist

    def imageWindow(self):
        file = self.sender().text()
        image(file)
        
    def videoWindow(self):
        file = self.sender().text()
        video(file)

    def center(self):
        frame_info = self.frameGeometry()
        display_center = QDesktopWidget().availableGeometry().center()
        frame_info.moveCenter(display_center)
        self.move(frame_info.topLeft())

    # 엑셀 추출 버튼 클릭 시
    def excelButtonClicked(self):
        if self.f_name == '':
            return

        #create Excel    
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "TongTong"
        col_excel = list(self.colname)[0:5]
        
        for x in range(1, len(col_excel) + 1):
            sheet.cell(row = 1, column = x).value = col_excel[x - 1]
            
        for x in range(0, len(self.rowlist)):
            for y in range(1, len(self.rowlist[x]) + 1):
                sheet.cell(row = x + 2, column = y).value = str(self.rowlist[x][y - 1])
                
        #resize the cell
        for x in range(0, len(self.colname)):
            MAX = 1
            for y in range(1, len(self.rowlist) + 1):
                cell_size = len(str(self.rowlist[y - 1][x]))
                if MAX < cell_size:
                    MAX = cell_size
                    sheet.column_dimensions[chr(65 + x)].width = MAX + 5
                sheet.row_dimensions[y].height = 20
        sheet.row_dimensions[y + 1].height = 20
        
        #change the font 
        for x in range(1, len(self.colname) + 1):
            cell = sheet[chr(64 + x) + "1"]
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(right=Side(border_style="thick"), bottom=Side(border_style="thick"))

        for x in range(len(self.rowlist)):
            for y in range(len(self.rowlist[x])):
                cell = sheet[chr(65 + y) + str(x + 2)]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(right=Side(border_style="thick"))
 
        wb.save("TongTong_" + self.f_name + ".xlsx")

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion'))
    path = 'C:/MDTool/SM-G955N/20210611-TongTong-001/TongTong/'
    ui = TongTongScreen(path)
    sys.exit(app.exec_())
