from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

from openpyxl.styles import Font, Border, Side, Alignment
import openpyxl

from exportDB import KaKaoTalk_DB_1, KaKaoTalk_DB_2
from button import Button
from videoWindow import video, image

import os
import copy
import urllib.request
class KakaoScreen(QDialog):
    def __init__(self, path):
        super().__init__()
        # 초기화
        self.setWindowFlags(Qt.WindowTitleHint | Qt.WindowCloseButtonHint)
        self.tableWidget = QTableWidget()
        self.on_off, self.f_name = 0, ''
        self.kakaoColnames, self.kakaoRowlists, self.kakao2Colnames, self.kakao2Rowlists = [], [], [], []

        self.path = path
        self.kakaoData()  # 미리 kakaotalk 데이터 모두 가져오기
        self.setupUI()

    def setupUI(self):

        # Window Backgrond
        palette = QPalette()
        palette.setColor(QPalette.Background, QColor(255, 255, 255))
        self.setAutoFillBackground(True)
        self.setPalette(palette)

        # Window Setting
        self.setGeometry(500, 70, 1200, 800)
        self.setWindowTitle("main")
        self.setFixedSize(self.rect().size())
        self.setContentsMargins(10, 10, 10, 10)

        # Ctrl+ F
        self.shortcut = QShortcut(QKeySequence('Ctrl+f'), self)
        self.shortcut.activated.connect(self.handleFind)

        # back/search button
        self.backButton = Button(QPixmap("image/back.png"), 35, self.showAppWindow)
        self.searchButton = Button(QPixmap("image/search.png"), 35, self.search_items)

        # 마우스 커서를 버튼 위에 올리면 모양 바꾸기
        self.backButton.setCursor(QCursor(Qt.PointingHandCursor))
        self.searchButton.setCursor(QCursor(Qt.PointingHandCursor))

        self.backButton.setToolTip('뒤로가기')
        self.searchButton.setToolTip('찾기 버튼\n단축키 : ctrl + f')

        # search text
        self.searchBox = QtWidgets.QLineEdit()
        self.searchBox.setMinimumSize(QtCore.QSize(0, 15))
        self.searchBox.returnPressed.connect(self.search_items)

        # excel button
        self.excelSaveButton = QPushButton(default=False, autoDefault=False)
        self.excelSaveButton.setFixedWidth(100)
        self.excelSaveButton.setText('Save as xls')
        self.excelSaveButton.clicked.connect(self.excelButtonClicked)
        self.excelSaveButton.setToolTip('현재 보고있는 표를 엑셀로 저장하는 버튼')

        # open combo box
        self.openComboBox = QComboBox()
        self.openComboBox.addItem("KakaoTalk.db")
        self.openComboBox.addItem("KakaoTalk2.db")
        self.openComboBox.setFixedWidth(100)
        self.openComboBox.activated.connect(self.DBClicked)
        self.openComboBox.setToolTip('KakaoTalk 데이터베이스')

        # combo box mes => db1
        self.kakaoComboBox = QComboBox()
        self.kakaoComboBox.addItem("chat_logs")
        self.kakaoComboBox.addItem("chat_rooms")
        self.kakaoComboBox.setFixedWidth(100)
        self.kakaoComboBox.activated.connect(self.kakaoComboEvent)
        self.kakaoComboBox.setToolTip('KakaoTalk.db의 tables')

        # combo box name => db2
        self.kakao2ComboBox = QComboBox()
        self.kakao2ComboBox.addItem("friends")
        self.kakao2ComboBox.setFixedWidth(100)
        self.kakao2ComboBox.activated.connect(self.kakao2ComboEvent)
        self.kakao2ComboBox.hide()
        self.kakao2ComboBox.setToolTip('KakaoTalk2.db의 tables')

        # combo chat room
        self.chatRoomComboBox = QComboBox()
        for i in range(self.chatRoomLen):
            self.chatRoomComboBox.addItem((self.chatRoomName[i]))
        self.chatRoomComboBox.activated.connect(self.chatRoomComboEvent)
        self.chatRoomComboBox.setToolTip('Chat Room')

        hbox1 = QHBoxLayout()
        hbox1.addWidget(self.backButton)
        hbox1.addStretch(1)
        hbox1.addWidget(self.searchBox)
        hbox1.addWidget(self.searchButton)
        hbox1.addWidget(self.openComboBox)
        hbox2 = QHBoxLayout()
        hbox2.addWidget(self.kakaoComboBox)
        hbox2.addWidget(self.kakao2ComboBox)
        hbox2.addWidget(self.chatRoomComboBox)
        hbox2.addStretch(1)
        hbox2.addWidget(self.excelSaveButton)
        layout = QVBoxLayout()
        layout.addLayout(hbox1)
        layout.addLayout(hbox2)
        layout.addWidget(self.tableWidget)

        self.setLayout(layout)
        self.center()
        self.show()

    def showAppWindow(self):
        self.close()

    # ctrl + f
    def handleFind(self):
        self.on_off = 1
        findDialog = QDialog()
        grid = QGridLayout()
        findDialog.setLayout(grid)
        findLabel = QLabel("Search...", findDialog)
        grid.addWidget(findLabel, 1, 0)
        self.findField = QLineEdit(findDialog)
        grid.addWidget(self.findField, 1, 1)
        findButton = QPushButton("Find", findDialog)
        findButton.clicked.connect(self.search_items)
        grid.addWidget(findButton, 2, 1)
        findDialog.setWindowTitle("Search items")
        findDialog.exec_()
        self.on_off = 0

    # search box
    def search_items(self):
        # rest font
        def reset(self, items):
            for item in items:
                if item == None:
                    pass
                else:
                    item.setBackground(QBrush(Qt.white))
                    item.setForeground(QBrush(Qt.black))
                    item.setFont(QFont())
        if self.on_off == 0:
            text = self.searchBox.text()
            selected_items = self.tableWidget.findItems(self.searchBox.text(), QtCore.Qt.MatchContains)
        else:
            text = self.findField.text()
            selected_items = self.tableWidget.findItems(self.findField.text(), QtCore.Qt.MatchContains)
        allitems = self.tableWidget.findItems("", QtCore.Qt.MatchContains)

        # reset
        reset(self, allitems)

        # highlight the search results
        for item in allitems:
            if item in selected_items:
                item.setBackground(QBrush(Qt.black))
                item.setForeground(QBrush(Qt.white))
                item.setFont(QFont("Helvetica", 9, QFont.Bold))

        if self.searchBox.text() == "" and self.findField.text() != "":
            pass
        elif self.searchBox.text() == "":
            reset(self, allitems)
            print("sb None")
        elif self.on_off == 1 and self.findField.text() == "":
            reset(self, allitems)
            print("ff None")

    # table combo select
    def kakaoComboEvent(self):
        if self.kakaoComboBox.currentText() == 'chat_logs':
            self.chatRoomComboBox.show()
            colname, rowlist = self.kakaoColnames[0], self.chatrowlists[0]
        elif self.kakaoComboBox.currentText() == 'chat_rooms':
            self.chatRoomComboBox.hide()
            colname, rowlist = self.kakaoColnames[1], self.kakaoRowlists[1]
        
        self.showTable(colname, rowlist)

    def kakao2ComboEvent(self):
        if self.kakao2ComboBox.currentText() == 'friends':
            colname, rowlist = self.kakao2Colnames[0], self.kakao2Rowlists[0]

        self.showTable(colname, rowlist)

    def chatroom(self):
        self.kakaoColnames[0] = list(self.kakaoColnames[0])
        self.kakaoColnames[0][2] = '받는사람'
        self.chatrowlists = []
        talkrowlist = self.kakaoRowlists[0]

        for k in range(self.chatRoomLen):
            crowlist = []
            for i in range(len(self.kakaoRowlists[0])):
                people = copy.deepcopy(self.chatRoomPeople[k])
                if talkrowlist[i][2] == self.chatRoomNum[k]:
                    if talkrowlist[i][1] in people:
                        people.remove(talkrowlist[i][1])
                    people=', '.join(people)
                    talkrowlist[i][2] = people
                    if people == '':
                        talkrowlist[i][2] = talkrowlist[i][1]
                    crowlist.append(talkrowlist[i])
            
            self.chatrowlists.append(crowlist)

    def chatRoomComboEvent(self):
        colname = self.kakaoColnames[0]
        for k in range(self.chatRoomLen):
            if self.chatRoomComboBox.currentText() == self.chatRoomName[k]:
                rowlist = self.chatrowlists[k]
        
        self.showTable(colname, rowlist)

    def kakaoData(self):
        self.kakao2Colnames, self.kakao2Rowlists = KaKaoTalk_DB_2(self.path)
        self.kakaoColnames, self.kakaoRowlists = KaKaoTalk_DB_1(self.path, self.kakao2Rowlists)
        
        for i in range(len(self.kakaoRowlists[0])):
            del self.kakaoRowlists[0][i][-1]
        self.kakaoColnames[0] = list(self.kakaoColnames[0])
        del self.kakaoColnames[0][-1]

        self.chatRoomLen = len(self.kakaoRowlists[1])
        self.chatRoomNum,self.chatRoomPeople,self.chatRoomName=[],[],[]
        for CR in self.kakaoRowlists[1]:
            self.chatRoomNum.append(CR[0])
            CS = CR[2].split(', ')
            self.chatRoomPeople.append(CS)
            if CR[1] == '':
                if len(CS) > 3: self.chatRoomName.append(', '.join(CS[:3]) + f' 외 {len(CS)}명')
                else: self.chatRoomName.append(CR[2])
            else: self.chatRoomName.append(CR[1])
        
        self.chatroom()
    
        colname, rowlist = self.kakaoColnames[0], self.chatrowlists[0]

        self.f_name = "KakaoTalk_db"
        self.showTable(colname, rowlist)

    # DB 버튼 클릭 시
    def DBClicked(self):

        if self.openComboBox.currentText() == 'KakaoTalk.db':
            self.kakao2ComboBox.hide()
            self.kakaoComboBox.show()
            self.chatRoomComboBox.show()
            self.kakaoComboBox.setCurrentIndex(0)
            colname, rowlist = self.kakaoColnames[0], self.kakaoRowlists[0]
            self.f_name = "KakaoTalk_db"

        elif self.openComboBox.currentText() == 'KakaoTalk2.db':
            self.kakaoComboBox.hide()
            self.kakao2ComboBox.show()
            self.chatRoomComboBox.hide()
            self.kakao2ComboBox.setCurrentIndex(0)
            colname, rowlist = self.kakao2Colnames[0], self.kakao2Rowlists[0]
            self.f_name = "KakaoTalk2_db"

        self.showTable(colname, rowlist)

    def tableHeaderClicked(self):
        # 헤더 click 시에만 정렬하고 다시 정렬기능 off
        # 정렬 계속 on 시켜 놓으면 다른 테이블 클릭 시 data 안보이는 현상 발생
        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.setSortingEnabled(False)

    def showTable(self, colname, rowlist):
        self.tableWidget.clear()

        self.tableWidget.setColumnCount(len(colname))  # col 개수 지정
        self.tableWidget.setRowCount(len(rowlist))  # row 개수 지정

        self.tableWidget.setHorizontalHeaderLabels(colname)  # 열 제목 지정
        self.tableHeader = self.tableWidget.horizontalHeader()
        self.tableHeader.sectionClicked.connect(self.tableHeaderClicked)

        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 표 수정 못하도록

        media, profile, types = -1, -1, -1
        for m in range(len(colname)):
            if list(colname)[m] == '파일':
                media = m
            if list(colname)[m] == '프로필 이미지':
                profile = m
            if list(colname)[m] == '타입':
                types = m
            
        # rowlist를 표에 지정하기
        for i in range(len(rowlist)):
            for j in range(len(rowlist[i])):
                if ((j == media and rowlist[i][types] != 'reply') or j == profile) and rowlist[i][j] != '':
                    if j == media: tp = rowlist[i][types]
                    else: tp = 'photo'

                    if tp == 'photo':
                        mpath = 'image/image.png'
                    elif tp == 'file sharing':
                        mpath = 'image/files.png'
                    elif tp == 'video':
                        mpath = 'image/video.png'
                    else:
                        mpath = 'image/noimage.png'
                    
                    self.btn1 = Button(QPixmap(mpath), 30, self.imageWindow)
                    self.btn1.setText(rowlist[i][j])
                    self.tableWidget.setCellWidget(i, j, self.btn1)
                else:  # 텍스트
                    item = QTableWidgetItem(str(rowlist[i][j]))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget.setItem(i, j, item)
        self.tableWidget.verticalHeader().setDefaultSectionSize(130)
        self.colname = colname
        self.rowlist = rowlist

    def imageWindow(self):
        imageUrl = self.sender().text()
        iname = imageUrl.split('/')[-1]
        mediaPath = self.path + 'kakaoMedia/' + iname
        try:
            if not os.path.exists(mediaPath):
                urllib.request.urlretrieve(imageUrl, mediaPath)
        except:
            mediaPath = 'image/noimage.png'
        image(mediaPath)

    def center(self):
        frame_info = self.frameGeometry()
        display_center = QDesktopWidget().availableGeometry().center()
        frame_info.moveCenter(display_center)
        self.move(frame_info.topLeft())

    # 엑셀 추출 버튼 클릭 시
    def excelButtonClicked(self):
        if self.f_name == '':
            return

        # create Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "KakaoTalk"
        col_excel = list(self.colname)[0:5]

        for x in range(1, len(col_excel) + 1):
            sheet.cell(row=1, column=x).value = col_excel[x - 1]

        for x in range(0, len(self.rowlist)):
            for y in range(1, len(self.rowlist[x]) + 1):
                sheet.cell(row=x + 2, column=y).value = str(self.rowlist[x][y - 1])

        # resize the cell
        for x in range(0, len(self.colname)):
            MAX = 1
            for y in range(1, len(self.rowlist) + 1):
                cell_size = len(str(self.rowlist[y - 1][x]))
                if MAX < cell_size:
                    MAX = cell_size
                    sheet.column_dimensions[chr(65 + x)].width = MAX + 5
                sheet.row_dimensions[y].height = 20
        sheet.row_dimensions[y + 1].height = 20

        # change the font
        for x in range(1, len(self.colname) + 1):
            cell = sheet[chr(64 + x) + "1"]
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(right=Side(border_style="thick"), bottom=Side(border_style="thick"))

        for x in range(len(self.rowlist)):
            for y in range(len(self.rowlist[x])):
                cell = sheet[chr(65 + y) + str(x + 2)]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(right=Side(border_style="thick"))

        wb.save("KakaoTalk_" + self.f_name + ".xlsx")


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion'))  # --> 없으면, 헤더색 변경 안됨.
    path = 'C:/MDTool/SM-G925S/20210613-KakaoTalk-1/KakaoTalk/'
    ui = KakaoScreen(path)
    sys.exit(app.exec_())