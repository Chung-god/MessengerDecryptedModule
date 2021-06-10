from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

from openpyxl.styles import Font, Border, Side, Alignment
import openpyxl
from xml.etree.ElementTree import parse

from exportDB import wickrDB
from button import Button

from videoWindow import video, image

import os
import copy

class WickrScreen(QDialog):
    def __init__(self, phoneNo):
        super().__init__()
        # 초기화
        self.setWindowFlags(Qt.WindowTitleHint | Qt.WindowCloseButtonHint)
        self.tableWidget = QTableWidget()
        self.on_off, self.f_name = 0, ''
        self.wickrColnames, self.wickrRowlists = [], []

        self.phoneNo = phoneNo
        self.path = f'C:/AppData/{self.phoneNo}/Wickr/'

        self.setupUI()

    def setupUI(self):

        # Window Backgrond
        palette = QPalette()
        palette.setColor(QPalette.Background, QColor(255, 255, 255))
        self.setAutoFillBackground(True)
        self.setPalette(palette)

        # Window Setting
        self.setGeometry(500, 70, 1200, 800)
        self.setWindowTitle("main")
        self.setFixedSize(self.rect().size())
        self.setContentsMargins(10, 10, 10, 10)

        # Ctrl+ F
        self.shortcut = QShortcut(QKeySequence('Ctrl+f'), self)
        self.shortcut.activated.connect(self.handleFind)

        # back/search button
        self.backButton = Button(QPixmap("image/back.png"), 45, self.showAppWindow)
        self.searchButton = Button(QPixmap("image/search.png"), 45, self.search_items)
        
        # 마우스 커서를 버튼 위에 올리면 모양 바꾸기
        self.backButton.setCursor(QCursor(Qt.PointingHandCursor))
        self.searchButton.setCursor(QCursor(Qt.PointingHandCursor))

        self.backButton.setToolTip('뒤로가기')
        self.searchButton.setToolTip('찾기 버튼\n단축키 : ctrl + f')

        # search text
        self.searchBox = QtWidgets.QLineEdit()
        self.searchBox.setMinimumSize(QtCore.QSize(0, 15))
        self.searchBox.returnPressed.connect(self.search_items)

        # excel button
        self.excelSaveButton = QPushButton(default=False, autoDefault=False)
        self.excelSaveButton.setFixedWidth(100)
        self.excelSaveButton.setText('Save as xls')
        self.excelSaveButton.clicked.connect(self.excelButtonClicked)
        self.excelSaveButton.setToolTip('현재 보고있는 표를 엑셀로 저장하는 버튼')

        # open combo box
        self.openComboBox = QComboBox()
        self.openComboBox.addItem("wickr.db")
        self.openComboBox.setFixedWidth(100)
        self.openComboBox.activated.connect(self.DBClicked)
        self.openComboBox.setToolTip('Wickr의 데이터베이스')

        # combo box wickr
        self.wickrComboBox = QComboBox()
        self.wickrComboBox.addItem("Wickr_Message")
        self.wickrComboBox.addItem("Wickr_User")
        self.wickrComboBox.addItem("Wickr_Convo")
        self.wickrComboBox.setFixedWidth(100)
        self.wickrComboBox.activated.connect(self.wickrComboEvent)
        self.wickrComboBox.setToolTip('wickr.db의 Tables')

        # combo chat room
        self.chatRoomComboBox = QComboBox()
        self.chatRoomComboBox.hide()

        self.password = QLabel()
        self.passerror = QLabel()
        
        hbox1 = QHBoxLayout()
        hbox1.addWidget(self.backButton)
        hbox1.addStretch(1)
        hbox1.addWidget(self.searchBox)
        hbox1.addWidget(self.searchButton)
        hbox1.addWidget(self.openComboBox)
        hbox2 = QHBoxLayout()
        hbox2.addWidget(self.wickrComboBox)
        hbox2.addWidget(self.chatRoomComboBox)
        hbox2.addStretch(1)
        hbox2.addWidget(self.passerror)
        hbox2.addStretch(1)
        hbox2.addWidget(self.excelSaveButton)

        layout = QVBoxLayout()
        layout.addLayout(hbox1)
        layout.addLayout(hbox2)
        layout.addWidget(self.tableWidget)

        self.setLayout(layout)
        self.center()
        self.show()

        self.showDialog()
        
        for i in range(self.chatRoomLen):
            self.chatRoomComboBox.addItem(self.chatRoomNum[i])
        self.chatRoomComboBox.activated.connect(self.chatRoomComboEvent)
        self.chatRoomComboBox.setToolTip('Chat Room')
        self.chatRoomComboBox.show()

    def showDialog(self):
        text, ok = QInputDialog.getText(self, 'Password', 'Enter your password :', QLineEdit.Password, flags=(Qt.WindowTitleHint|Qt.WindowCloseButtonHint))
        
        if ok:
            self.password.setText(str(text))
            self.wickrData()  # 미리 Wickr 데이터 모두 가져오기
        else:
            self.passerror.setText('데이터를 가져올 수 없습니다.')

    def showAppWindow(self):
        self.close()

    # ctrl + f
    def handleFind(self):
        self.on_off = 1
        findDialog = QDialog()
        grid = QGridLayout()
        findDialog.setLayout(grid)
        findLabel = QLabel("Search...", findDialog)
        grid.addWidget(findLabel, 1, 0)
        self.findField = QLineEdit(findDialog)
        grid.addWidget(self.findField, 1, 1)
        findButton = QPushButton("Find", findDialog)
        findButton.clicked.connect(self.search_items)
        grid.addWidget(findButton, 2, 1)
        findDialog.setWindowTitle("Search items")
        findDialog.exec_()
        self.on_off = 0

    # search box
    def search_items(self):

        # rest font
        def reset(self, items):
            for item in items:
                if item == None:
                    pass
                else:
                    item.setBackground(QBrush(Qt.white))
                    item.setForeground(QBrush(Qt.black))
                    item.setFont(QFont())

        if self.on_off == 0:
            text = self.searchBox.text()
            selected_items = self.tableWidget.findItems(self.searchBox.text(), QtCore.Qt.MatchContains)
        else:
            text = self.findField.text()
            selected_items = self.tableWidget.findItems(self.findField.text(), QtCore.Qt.MatchContains)

        allitems = self.tableWidget.findItems("", QtCore.Qt.MatchContains)

        # reset
        reset(self, allitems)

        # highlight the search results
        for item in allitems:
            if item in selected_items:
                item.setBackground(QBrush(Qt.black))
                item.setForeground(QBrush(Qt.white))
                item.setFont(QFont("Helvetica", 9, QFont.Bold))

        if self.searchBox.text() == "" and self.on_off == 0:
            reset(self, allitems)
            pass

        elif self.searchBox.text() == "":
            reset(self, allitems)
            print("sb None")
        elif self.on_off == 1 and self.findField.text() == "":
            reset(self, allitems)
            print("ff None")

    # table combo select
    def wickrComboEvent(self):
        if self.wickrComboBox.currentText() == 'Wickr_Message':
            self.chatRoomComboBox.show()
            colname, rowlist = self.wickrColnames[0], self.chatrowlists[0]
        elif self.wickrComboBox.currentText() == 'Wickr_User':
            self.chatRoomComboBox.hide()
            colname, rowlist = self.wickrColnames[1], self.wickrRowlists[1]
        elif self.wickrComboBox.currentText() == 'Wickr_Convo':
            self.chatRoomComboBox.hide()
            colname, rowlist = self.wickrColnames[2], self.wickrRowlists[2]

        self.showTable(colname, rowlist)

    def chatroom(self):
        self.wickrColnames[0] = list(self.wickrColnames[0])
        self.wickrColnames[0][2] = '받는사람'
        self.chatrowlists = []
        talkrowlist = self.wickrRowlists[0]

        for k in range(self.chatRoomLen):
            crowlist = []
            for i in range(len(self.wickrRowlists[0])):
                people = copy.deepcopy(self.chatRoomPeople[k])
                if talkrowlist[i][2] == self.chatRoomNum[k]:
                    if talkrowlist[i][1] in people:
                        people.remove(talkrowlist[i][1])
                    people=', '.join(people)
                    talkrowlist[i][2] = people
                    crowlist.append(talkrowlist[i])
            
            self.chatrowlists.append(crowlist)
    
    def chatRoomComboEvent(self):
        
        colname = self.wickrColnames[0]
        for k in range(self.chatRoomLen):
            if self.chatRoomComboBox.currentText() == self.chatRoomNum[k]:
                rowlist = self.chatrowlists[k]
        
        self.showTable(colname, rowlist)

    def wickrData(self):
        try:
            self.wickrColnames, self.wickrRowlists = wickrDB(self.path, self.password.text())
        
            self.chatRoomLen = len(self.wickrRowlists[2])
            self.chatRoomNum = [self.wickrRowlists[2][i][0] for i in range(self.chatRoomLen)]
            self.chatRoomPeople = [self.wickrRowlists[2][i][1].split(', ') for i in range(self.chatRoomLen)]

            self.chatroom()

            colname, rowlist = self.wickrColnames[0], self.chatrowlists[0]

            self.f_name = "wickr_db"
            self.showTable(colname, rowlist)
            self.passerror.hide()
        except:
            self.passerror.setText('비밀번호가 틀렸습니다.')
            self.showDialog()

    # DB 버튼 클릭 시
    def DBClicked(self):

        if self.openComboBox.currentText() == 'wickr.db':
            self.wickrComboBox.setCurrentIndex(0)
            colname, rowlist = self.wickrColnames[0], self.wickrRowlists[0]
            self.f_name = "wickr_db"

        self.showTable(colname, rowlist)

    def tableHeaderClicked(self):
        # 헤더 click 시에만 정렬하고 다시 정렬기능 off
        # 정렬 계속 on 시켜 놓으면 다른 테이블 클릭 시 data 안보이는 현상 발생
        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.setSortingEnabled(False)

    def showTable(self, colname, rowlist):
        self.tableWidget.clear()

        self.tableWidget.setColumnCount(len(colname))  # col 개수 지정
        self.tableWidget.setRowCount(len(rowlist))  # row 개수 지정

        self.tableWidget.setHorizontalHeaderLabels(colname)  # 열 제목 지정
        self.tableHeader = self.tableWidget.horizontalHeader()
        self.tableHeader.sectionClicked.connect(self.tableHeaderClicked)

        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 표 너비 지정
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 표 수정 못하도록
        
        media = -1
        for m in range(len(colname)):
            if list(colname)[m] == '미디어':
                media = m
            elif list(colname)[m] == '타입':
                types = m
                
        # rowlist를 표에 지정하기
        for i in range(len(rowlist)):
            for j in range(len(rowlist[i])):
                if j == media:
                    tp = rowlist[i][types][:5]
                    if tp == 'image': # 사진
                        if os.path.isfile(self.path+'files/dec/'+rowlist[i][j]) == False: #원본파일 없을 경우
                            mpath = 'image/noimage.png'
                        else: # 원본 파일 있는 경우
                            mpath = 'image/image.png'
                        self.btn1 = Button(QPixmap(mpath), 30, self.imageWindow)
                        self.btn1.setText(rowlist[i][j])
                        self.btn1.setCursor(QCursor(Qt.PointingHandCursor))
                        self.tableWidget.setCellWidget(i,j,self.btn1)

                    elif tp == 'video': # 비디오
                        if os.path.isfile(self.path+'files/dec/'+rowlist[i][j]) == False: #원본파일 없을 경우
                            mpath = 'image/noimage.png'
                        else: # 원본 파일 있는 경우
                            mpath = 'image/video.png'
                        self.btn2 = Button(QPixmap(mpath), 30, self.videoWindow)
                        self.btn2.setText(rowlist[i][j])
                        self.btn2.setCursor(QCursor(Qt.PointingHandCursor))
                        self.tableWidget.setCellWidget(i,j,self.btn2)
                else: # 텍스트
                    item = QTableWidgetItem(str(rowlist[i][j]))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget.setItem(i, j, item)

        self.tableWidget.verticalHeader().setDefaultSectionSize(130)
        self.colname = colname
        self.rowlist = rowlist

    def imageWindow(self):
        file = self.sender().text()
        mediaPath = self.path+'files/dec/'+file
        image(mediaPath)

    def videoWindow(self):
        file = self.sender().text()
        mediaPath = self.path+'files/dec/'+file
        video(mediaPath)
    
    def center(self):
        frame_info = self.frameGeometry()
        display_center = QDesktopWidget().availableGeometry().center()
        frame_info.moveCenter(display_center)
        self.move(frame_info.topLeft())

    # 엑셀 추출 버튼 클릭 시
    def excelButtonClicked(self):
        if self.f_name == '':
            return

        # create Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Wickr"
        col_excel = list(self.colname)[0:5]

        for x in range(1, len(col_excel) + 1):
            sheet.cell(row=1, column=x).value = col_excel[x - 1]

        for x in range(0, len(self.rowlist)):
            for y in range(1, len(self.rowlist[x]) + 1):
                sheet.cell(row=x + 2, column=y).value = str(self.rowlist[x][y - 1])

        # resize the cell
        for x in range(0, len(self.colname)):
            MAX = 1
            for y in range(1, len(self.rowlist) + 1):
                cell_size = len(str(self.rowlist[y - 1][x]))
                if MAX < cell_size:
                    MAX = cell_size
                    sheet.column_dimensions[chr(65 + x)].width = MAX + 5
                sheet.row_dimensions[y].height = 20
        sheet.row_dimensions[y + 1].height = 20

        # change the font
        for x in range(1, len(self.colname) + 1):
            cell = sheet[chr(64 + x) + "1"]
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(right=Side(border_style="thick"), bottom=Side(border_style="thick"))

        for x in range(len(self.rowlist)):
            for y in range(len(self.rowlist[x])):
                cell = sheet[chr(65 + y) + str(x + 2)]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(right=Side(border_style="thick"))

        wb.save("Wickr_" + self.f_name + ".xlsx")


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion')) # --> 없으면, 헤더색 변경 안됨.
    phoneNo = 'SM-G955N'
    ui = WickrScreen(phoneNo)
    sys.exit(app.exec_())
